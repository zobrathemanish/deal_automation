import os
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
# from geopy.geocoders import Nominatim
# from geopy.exc import GeocoderTimedOut
import re
from openpyxl.utils.datetime import from_excel
import datetime


# Configure Flask App
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {"xlsx", "xlsm"}

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Ensure upload directory exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    """ Check if the uploaded file has an allowed extension. """
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# Function to split the address into unit number, street number, name, and route type
def split_address(address):
    if address:
        match = re.match(r"(?:Unit\s*(\d+)\s*-\s*)?(\d+)\s+(.+)\s+(\w+)$", address)
        if match:
            unit_number = match.group(1) if match.group(1) else None  # Extract Unit Number if exists
            street_number = match.group(2)
            street_name = match.group(3)
            route_type = match.group(4)
            return unit_number, street_number, street_name, route_type
    return None, None, None, None

def format_date(value):
    """
    Converts an Excel datetime value to a string in DD/MM/YYYY format.
    If the value is None or not a datetime, returns it as-is.
    """
    if isinstance(value, datetime.datetime):
        return value.strftime("%d/%m/%Y")  # Convert datetime to "DD/MM/YYYY"
    return value  # Return as-is if not a date

def format_currency(value):
    """
    Converts a numeric value to a currency format ($xx.xx).
    If the value is None or not a number, returns it as-is.
    """
    if isinstance(value, (int, float)):
        return f"${value:,.2f}"  # Formats as "$12.00"
    return value  # Return as-is if not a valid number
def extract_rate(value):
    """
    Extracts 'Net' or 'Gross' from a string like 'Net Rate' or 'Gross Rate'.
    If the value is not a valid string, return it as-is.
    """
    if isinstance(value, str):
        return value.split()[0]  # Takes the first word (Net or Gross)
    return value  # Return as-is if not a string


def extract_data_lease_from_excel(file_path):
    """
    Extracts relevant data from the uploaded Excel file (Deal Sheet).
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active  # Get the active worksheet
    address = ws["C29"].value
    unit_number, street_number, street_name, route_type = split_address(address)


    
    extracted_data_lease = {
        "File Number": ws["H7"].value,
        "Unit Number": unit_number,
        "Street Number": street_number,
        "Street Name": street_name,
        "Route Type": route_type,
        "Year Built": ws["G91"].value,
        "Building Size (sf)": ws["G92"].value,
        "Single/Multi-Tenant": ws["G90"].value,
        "Lease Start Date": format_date(ws["G87"].value),
        "Lease End Date": format_date(ws["G89"].value),
        "Term Length (Years)": ws["G88"].value / 12 if ws["G88"].value else None,
        "Unit Size (SF)": ws["G86"].value,
        "Initial Net Rate (PSF)": format_currency(ws["E38"].value),
        "Tenant Name": ws["B15"].value,
        "Property Type": ws["D86"].value,
        "Leasehold Improvements & Other Incentives (TI)": ws["D90"].value if ws["D90"].value else "$0.00",
        "Net Effective Rate": ws["D92"].value,
        "Type of Lease": ws["B3"].value,
        "Rate": extract_rate(ws["E3"].value),
        "Total Operating Expenses": format_currency(ws["D88"].value),
        "Client": ws["B5"].value,
        "Lessor_Company": ws["B7"].value,
        "Lessor_Contact_Name": ws["B8"].value,
        "Lessor_Address": ws["B9"].value,
        "Lessor_Phone_Number": ws["B10"].value,
        "Lessor_Email": ws["B11"].value,
        "Lessor_City": ws["B12"].value,
        "Lessor_Province": ws["B13"].value,
        "Lessor_Postal_Code": ws["E13"].value,
        "Lessee_Company": ws["B15"].value,
        "Lessee_Contact_Name": ws["B16"].value,
        "Lessee_Address": ws["B17"].value,
        "Lessee_Phone_Number": ws["B18"].value,
        "Lessee_Email": ws["B19"].value,
        "Lessee_City": ws["B20"].value,
        "Lessee_Province": ws["B21"].value,
        "Lessee_Postal_Code": ws["E21"].value,
    }

      # Extract non-empty "Source/Agent" values from D71 to D80
    source_agent_values = [ws[f"D{i}"].value for i in range(71, 81) if ws[f"D{i}"].value]
    extracted_data_lease["Source/ Agent"] = ", ".join(source_agent_values)

    # Extract comments from A38 to A47 and E38 to E47
    comments = []
    for i in range(38, 48):
        comment_key = ws[f"A{i}"].value  # Term length (in months)
        comment_value = ws[f"E{i}"].value  # Rate ($ value)

        # Ensure comment_key (months) is a valid integer
        if isinstance(comment_key, (int, float)):
            comment_key = f"{int(comment_key)} months"  # Convert to "36 months", "24 months", etc.

        # Ensure comment_value (rate) is formatted correctly as "$xx.xx"
        if isinstance(comment_value, (int, float)):
            comment_value = f"${comment_value:,.2f}"  # Converts 24 to "$24.00"

        # Add only if either key or value exists
        if comment_key and comment_value:
            comments.append(f"{comment_key}@{comment_value}")

    # Join all comments into a single string
    extracted_data_lease["Comments"] = ", ".join(comments)

    return extracted_data_lease

def extract_data_retail_sales_from_excel(file_path):
    """
    Extracts relevant data from the uploaded Excel file (Deal Sheet).
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active  # Get the active worksheet
    # Extract address and split it
    address = ws["C39"].value
    unit_number, street_number, street_name, route_type = split_address(address)
    
    extracted_data_retail_sales = {
        "Unit Number": unit_number,
        "Street Name": street_name,
        "Route Type": route_type,
        "Year Built": ws["G91"].value,       
        "Sold Price": ws["C45"].value,
        "Sale Date": format_date(ws["C70"].value),
        "Price PSF": format_currency(ws["F72"].value),
        "Client": ws["B3"].value,
        "Vendor_Company": ws["B5"].value,
        "Vendor_Contact_Name": ws["B6"].value,
        "Vendor_Address": ws["B7"].value,
        "Vendor_Phone_Number": ws["B8"].value,
        "Vendor_Email": ws["B9"].value,
        "Vendor_City": ws["B10"].value,
        "Vendor_Province": ws["B11"].value,
        "Vendor_Postal_Code": ws["E11"].value,
        "Purchaser_Company": ws["B19"].value,
        "Purchaser_Contact_Name": ws["B20"].value,
        "Purchaser_Address": ws["B21"].value,
        "Purchaser_Phone_Number": ws["B22"].value,
        "Purchaser_Email": ws["B23"].value,
        "Purchaser_City": ws["B24"].value,
        "Purchaser_Province": ws["B25"].value,
        "Purchaser_Postal_Code": ws["E25"].value,
    }
      

    return extracted_data_retail_sales

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        data_type = request.form.get("data_type")

        if data_type == "lease":
            # Process File Upload
            file = request.files["file"]
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(file_path)

                # Extract Data from Excel
                excel_data = extract_data_lease_from_excel(file_path)
            else:
                excel_data = {}  # No file uploaded, empty dictionary

            # Get Data from Form
            form_data = {
            "Province": request.form.get("lease_province"),
            "City": request.form.get("lease_city"),
            "Quadrant": request.form.get("lease_quadrant"),
            "Submarket": request.form.get("lease_submarket"),
            "Land Actual sf": request.form.get("lease_land_actual_sf"),
            "Zoning": request.form.get("lease_zoning"),
            "Parcel Use Code": request.form.get("lease_parcel_use_code"),
            "Gross Lease Rate (PSF)": request.form.get("lease_gross_lease_rate"),
            "Data Comments": request.form.get("lease_data_comments"),  # FIXED
            "CAM": request.form.get("lease_cam"),
            "Property Taxes": request.form.get("lease_property_taxes"),
            "Roll Number": request.form.get("lease_roll_no")
}


            # Combine Excel and Form Data
            final_data = {**excel_data, **form_data}

            # Define final column order
            column_order = [
                "File Number", "Unit Number", "Street Number", "Street Name", "Route Type",
                "Municipality", "Province", "Quadrant", "Submarket", "Land Actual sf",
                "Zoning", "Year Built", "Building Size (sf)", "Single/Multi-Tenant",
                "Parcel Use Code", "Lease Start Date", "Lease End Date", "Term Length (Years)",
                "Unit Size (SF)", "Initial Net Rate (PSF)", "Gross Lease Rate (PSF)",
                "Tenant Name", "Property Type", "Source/ Agent", "Comments",
                "Leasehold Improvements & Other Incentives (TI)", "Net Effective Rate",
                "Type of Lease", "Rate", "Total Operating Expenses", "CAM", "Property Taxes",
                "Data Comments", "Updated in MySQL DB", "Roll Number", "Client",
                "Lessor_Company", "Lessor_Contact_Name", "Lessor_Address",
                "Lessor_Phone_Number", "Lessor_Email", "Lessor_City", "Lessor_Province",
                "Lessor_Postal_Code", "Lessee_Company", "Lessee_Contact_Name",
                "Lessee_Address", "Lessee_Phone_Number", "Lessee_Email", "Lessee_City",
                "Lessee_Province", "Lessee_Postal_Code", "Sabre_Aligned"
            ]

            # Create a DataFrame and ensure the correct order
            df = pd.DataFrame([final_data])
            df = df.reindex(columns=column_order)

            # Save to Excel
            output_path = os.path.join(UPLOAD_FOLDER, "data_record.xlsx")
            df.to_excel(output_path, index=False)

            return render_template("result.html", data_type=data_type, data=final_data, file_path=output_path)
        
        if data_type == "retail":
            # Process File Upload
            file = request.files["file"]
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(file_path)

                # Extract Data from Excel
                excel_data = extract_data_retail_sales_from_excel(file_path)
            else:
                excel_data = {}  # No file uploaded, empty dictionary

            # Get Data from Form
            form_data = {
                "City": request.form.get("retail_city"),
                "Quadrant": request.form.get("retail_quadrant"),
                "Submarket": request.form.get("retail_submarket"),
                "Land Actual sf": request.form.get("retail_land_actual_sf"),
                "Zoning": request.form.get("retail_zoning"),
                "Parcel Use Code": request.form.get("retail_parcel_use_code"),
                "Year Built": request.form.get("retail_year_built"),
                "Data Comments": request.form.get("retail_data_comments"),
                "Gross Upper Area": request.form.get("retail_gross_upper_area"),
                "Plan Area": request.form.get("retail_plan_area"),
                "Roll Number": request.form.get("retail_roll_no"),
                "Single/Multi-Tenant": request.form.get("retail_single_multi_tenant"),
                "Instrument Number": request.form.get("retail_instrument_number"),
                "Source": request.form.get("retail_source"),
                "Data Comments": request.form.get("retail_data_comments")
            }

            # Combine Excel and Form Data
            final_data = {**excel_data, **form_data}

            # Define final column order for Retail Sales Data
            column_order = [
                "Index #","Street Name", "Route Type", "Municipality", "Land Actual sf",
                "Zoning", "Year Built", "Gross Upper Area", "Plan Area", "Sold Price", "Sale Date", 
                "Price PSF", "Quadrant", "Submarket", "Single / Multi-Tenant", "Par Use Code", 
                "Instrument #", "Roll Number", "Source", "Comments", "Data Comments", 
                "Updated in MySQL DB", "Client", "Vendor_Company", "Vendor_Contact_Name", 
                "Vendor_Address", "Vendor_Phone_Number", "Vendor_Email", "Vendor_City", 
                "Vendor_Province", "Vendor_Postal_Code", "Purchaser_Company", "Purchaser_Contact_Name", 
                "Purchaser_Address", "Purchaser_Phone_Number", "Purchaser_Email", "Purchaser_City", 
                "Purchaser_Province", "Purchaser_Postal_Code"
            ]


            # Create a DataFrame and ensure the correct order
            df = pd.DataFrame([final_data])
            df = df.reindex(columns=column_order)

            # Save to Excel
            output_path = os.path.join(UPLOAD_FOLDER, "data_record.xlsx")
            df.to_excel(output_path, index=False)

            return render_template("result.html", data_type=data_type, data=final_data, file_path=output_path)

    return render_template("index.html")

from flask import send_file

@app.route("/download")
def download_file():
    """Serves the generated Excel file for download."""
    file_path = os.path.join(app.config["UPLOAD_FOLDER"], "data_record.xlsx")

    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "File not found", 404


if __name__ == "__main__":
    app.run(debug=True)
